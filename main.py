from datetime import datetime
from getpass import getpass
from dotenv import load_dotenv
import os
from jinja2 import Environment, FileSystemLoader
import openpyxl
import glob


# Create python dictionary from xlsx file
def get_config_items_xlsx(
    file_switch_variables, admin_pass, snmp_user_pass, radius_key
):
    try:
        workbook = openpyxl.load_workbook(file_switch_variables)

        config_items = {}

        # add timestamp
        config_items["time_now"] = get_time()

        # add all base config
        sheet = workbook["base"]
        config_items["hostname"] = sheet.cell(row=1, column=2).value
        config_items["software_file"] = sheet.cell(row=2, column=2).value
        config_items["snmp_location"] = sheet.cell(row=3, column=2).value
        config_items["default_gateway"] = sheet.cell(row=4, column=2).value
        config_items["source_vlan"] = sheet.cell(row=5, column=2).value
        config_items["ap_vlan"] = sheet.cell(row=6, column=2).value
        config_items["ap_wlan"] = sheet.cell(row=7, column=2).value

        # add all secret config
        config_items["admin_pass"] = admin_pass
        config_items["snmp_user_pass"] = snmp_user_pass
        config_items["radius_key"] = radius_key

        # add all vlan config
        sheet = workbook["vlan"]
        config_items["vlans"] = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True
        ):
            if row[0] != "vlan id":
                config_items["vlans"].append({"id": row[0], "name": row[1]})

        # add all ip config
        sheet = workbook["ip"]
        config_items["ip_int"] = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True
        ):
            if row[0] != "vlan":
                config_items["ip_int"].append(
                    {"vlan": row[0], "ip": row[1], "mask": row[2]}
                )

        # add all interface config
        sheet = workbook["interfaces"]
        config_items["interfaces"] = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=5, values_only=True
        ):
            if row[0] != "interface":
                config_items["interfaces"].append(
                    {
                        "interface": row[0],
                        "description": row[1],
                        "vlan": row[2],
                        "type": row[3],
                        "security": row[4],
                    }
                )

        return config_items

    except FileNotFoundError:
        print(f"The file {file_switch_variables} does not exist.")
    except PermissionError:
        print(
            f"\nPermission denied when trying to open {file_switch_variables}. \nPlease close the xlsx file so the script can access it."
        )
    except Exception as e:
        print(f"An error occurred when trying to open {file_switch_variables}: {e}")


# Get the current local date/time and format the object to a string in a readable format
def get_time():
    time = datetime.now()
    time = time.strftime("%Y-%m-%d_%Hh%Mm")
    return time


if __name__ == "__main__":
    # Check if directories for xlsx files and generated switch config exist, if not create them
    os.makedirs("switch_variables/", exist_ok=True)
    os.makedirs("generated_configs/", exist_ok=True)

    # Load environment variables, store secrets if present, otherwise prompt for input
    load_dotenv()
    admin_pass = (
        os.getenv("ADMIN_PASS")
        if os.getenv("ADMIN_PASS")
        else getpass(prompt="\nInput the Admin password:")
    )
    snmp_user_pass = (
        os.getenv("SNMP_USER_PASS")
        if os.getenv("SNMP_USER_PASS")
        else getpass(prompt="\nInput the SNMP user password:")
    )
    radius_key = (
        os.getenv("RADIUS_KEY")
        if os.getenv("RADIUS_KEY")
        else getpass(prompt="\nInput the Radius key:")
    )

    # Set location to folder with all xlsx files
    folder_switch_variables = "switch_variables"

    # Get lcoation all xlsx files in folder
    xlsx_files = glob.glob(os.path.join(folder_switch_variables, "*.xlsx"))

    # Loop over all xlsx files
    for file_path in xlsx_files:

        # Create python dictionary from xlsx file and given secrets
        config_itmes = get_config_items_xlsx(
            file_path, admin_pass, snmp_user_pass, radius_key
        )

        # Create jinja environment and load the template file
        env = Environment(
            loader=FileSystemLoader("."), trim_blocks=True, lstrip_blocks=True
        )
        template = env.get_template("full_config_template.j2")

        # Generate config by rendering the dictionary that was created from the xlsx file, after create txt file with hostname switch as name
        if config_itmes != None:
            hostname = f"generated_configs/{config_itmes.get('hostname')}.txt"
            cisco_config = template.render(config_itmes)
            with open(hostname, "w") as w:
                w.write(cisco_config)
                print()
                print("*" * 60)
                print(
                    f"Config file for {config_itmes.get('hostname')} has been generated."
                )
                print("*" * 60)
